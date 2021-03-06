# author: Vigneshwar Padmanaban
# date created: Dec 09, 2017

import openpyxl
import os
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
# Todo: Handle bad inputs everywhere
class AutoExcel:
    targetInputFile = None
    wb = None
    sh1 = None
    wb_O = None
    sh1_O = None
    sh2_O = None
    sh3_O = None
    groupLength = []
    columnStart = 'A'
    columnStart3 = 'A'
    index = 1
    sh3_index = 1
    prevNegative = 'N'
    potential_col = None
    current_col = None
    time_col = None
    capacity_col = None
    chdc_gap = 0
    rowStarts = 0
    rowStart = 0
    version = 0.1
    choice = None
    listOfSheets = []
    maxRows = 0
    outputFile = 'dummy'

    @staticmethod
    def loadSheets():
        for the_file in os.listdir('uploads'):
            file_path = os.path.join('uploads', the_file)
            try:
                if the_file.endswith('xlsx'):
                    AutoExcel.targetInputFile = file_path
            except Exception as e:
                print(e)
        print(AutoExcel.targetInputFile)

        # Input Excel
        print("Loading Input file: " + AutoExcel.targetInputFile)
        AutoExcel.wb = openpyxl.load_workbook(AutoExcel.targetInputFile)
        # Get Correct Sheet
        AutoExcel.listOfSheets = AutoExcel.wb.get_sheet_names()
        return AutoExcel.listOfSheets
    @staticmethod
    def initialize(valueList):

        AutoExcel.sh1 = AutoExcel.wb.get_sheet_by_name(valueList[0])
        AutoExcel.maxRows = AutoExcel.sh1.max_row
        # Output Excel
        AutoExcel.wb_O = openpyxl.Workbook()
        AutoExcel.sh1_O = AutoExcel.wb_O.create_sheet(index=0, title='-Current')
        AutoExcel.sh2_O = AutoExcel.wb_O.create_sheet(index=1, title='+Current')
        AutoExcel.sh3_O = AutoExcel.wb_O.create_sheet(index=2, title='Charge_Discharge')
        AutoExcel.outputFile = 'output'
        AutoExcel.rowStart = int(valueList[5])
        AutoExcel.potential_col = valueList[1]
        AutoExcel.current_col = valueList[2]
        AutoExcel.time_col = valueList[4]
        AutoExcel.capacity_col = valueList[3]
        AutoExcel.chdc_gap = int(valueList[6])
        print('xxx Initialized xxx')

    @staticmethod
    def negativeProc(current,i):
        AutoExcel.prevNegative = 'Y'
        potentialCoordinate = AutoExcel.columnStart + str(AutoExcel.index)
        currentCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 1) + str(AutoExcel.index)
        timeCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 2) + str(AutoExcel.index)
        capacityCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 3) + str(AutoExcel.index)

        # print("the coordinates are: " + potentialCoordinate + " " + currentCoord + " " + capacityCoord)
        potential = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.potential_col))
        time = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.time_col))
        capacity = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.capacity_col))

        AutoExcel.sh1_O[potentialCoordinate] = potential.value
        AutoExcel.sh1_O[currentCoord] = current.value
        AutoExcel.sh1_O[timeCoord] = time.value
        AutoExcel.sh1_O[capacityCoord] = capacity.value
        AutoExcel.index += 1
        # print("Row#: " + str(i) + "  Saving " + str(current.value) + " in output..")

        if i == AutoExcel.maxRows:
            AutoExcel.groupLength.append(AutoExcel.index - 1)
            AutoExcel.capture_Ch_Dc(i+1)

    @staticmethod
    def capture_Ch_Dc(i):
        if AutoExcel.chdc_gap==None or AutoExcel.chdc_gap==" " or AutoExcel.chdc_gap=="/n":
            AutoExcel.chdc_gap = 0
        lastIndex = i-1-int(AutoExcel.chdc_gap)
        potentialCoordinate3 = AutoExcel.columnStart3 + str(AutoExcel.sh3_index)
        currentCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 1) + str(AutoExcel.sh3_index)
        timeCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 2) + str(AutoExcel.sh3_index)
        capacityCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 3) + str(AutoExcel.sh3_index)
        print("the coordinates are: " + potentialCoordinate3 + " " + currentCoord3 + " " +timeCoord3 + " "+capacityCoord3)

        potential3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.potential_col))
        current3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.current_col))
        time3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.time_col))
        capacity3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.capacity_col))

        AutoExcel.sh3_O[potentialCoordinate3] = potential3.value
        AutoExcel.sh3_O[currentCoord3] = current3.value
        AutoExcel.sh3_O[timeCoord3] = time3.value
        AutoExcel.sh3_O[capacityCoord3] = capacity3.value
        AutoExcel.sh3_index += 1

    @staticmethod
    def addChart(groupLength1):
        print('Printing Chart...')
        # print('grouplength1: '+str(len(groupLength1)))
        chart = ScatterChart()
        chart.title = "Capacity Vs Voltage"
        chart.style = 13
        chart.x_axis.title = 'Capacity'
        chart.y_axis.title = 'Voltage'
        le = (len(groupLength1) * 4) + len(groupLength1)
        # print("The sizes of each grp" + str(groupLength1[0]))
        k = 0
        m = 1
        # print("list siz = " + str(len(groupLength1)))
        # print("le = " + str(le))
        for j in range(4, le, 5):
            print("j = " + str(j))
            xvalues = Reference(AutoExcel.sh1_O, min_col=j, min_row=1, max_row=groupLength1[k])
            values = Reference(AutoExcel.sh1_O, min_col=m, min_row=1, max_row=groupLength1[k])
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            k += 1
            m += 5

        AutoExcel.sh1_O.add_chart(chart, 'L3')

        out_f = "C:\\Users\\Vignesh\\PycharmProjects\\AutoExcel\\src\\uploads\\" + AutoExcel.outputFile+'.xlsx'
        AutoExcel.wb_O.save(out_f)

    @staticmethod
    def mainApp():
        AutoExcel.groupLength = []
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        for i in range(AutoExcel.rowStart, AutoExcel.maxRows+1):
            current = AutoExcel.sh1.cell(row=i,column=column_index_from_string(AutoExcel.current_col))
            if type(AutoExcel.columnStart) == type(current.value):
                currentVal = float(current.value)
            elif type(AutoExcel.index) == type(current.value) or type(AutoExcel.version) == type(current.value):
                currentVal = current.value
            else:
                pass
            if current.value == None:
                current.value = 0
            if currentVal<0:
                if AutoExcel.prevNegative == 'N':
                   AutoExcel.capture_Ch_Dc(i)
                AutoExcel.negativeProc(current,i)
            elif AutoExcel.prevNegative == 'Y':
                AutoExcel.capture_Ch_Dc(i)
                AutoExcel.prevNegative = 'N'
                potentialCoordinate = AutoExcel.columnStart + str(AutoExcel.index)
                currentCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 1) + str(AutoExcel.index)
                timeCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 2) + str(AutoExcel.index)
                capacityCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 3) + str(AutoExcel.index)
                AutoExcel.sh1_O[currentCoord].fill = redFill
                AutoExcel.sh1_O[potentialCoordinate].fill = redFill
                AutoExcel.sh1_O[timeCoord].fill = redFill
                AutoExcel.sh1_O[capacityCoord].fill = redFill
                AutoExcel.groupLength.append(AutoExcel.index-1)
                AutoExcel.index = 1
                AutoExcel.columnStart = get_column_letter(int(column_index_from_string(AutoExcel.columnStart))+5)
            else:
                continue

        print("Saving values..")
        out_f = "C:\\Users\\Vignesh\\PycharmProjects\\AutoExcel\\src\\uploads\\" + AutoExcel.outputFile + '.xlsx'
        AutoExcel.wb_O.save(out_f)
        AutoExcel.addChart(AutoExcel.groupLength)
        print("xxxxx------ Program Ended ------xxx")
        print("Output File Created: "+AutoExcel.outputFile+".xlsx")