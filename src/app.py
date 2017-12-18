# author: Vigneshwar Padmanaban
# date created: Dec 09, 2017

from sys import argv
import openpyxl
import sys
from openpyxl.styles import Color, PatternFill, Font, Border
# from openpyxl.styles import colors
# from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
import pyexcel as p

class AutoExcel:
    #Check Excel format
    fileName = argv[1]
    splitName = fileName.split('.')
    splitName[1] = fileName[-3:]
    if(splitName[1] == 'xls'):
        print('Its a xls file. Converting to .xlsx file..')
        p.save_book_as(file_name=fileName,dest_file_name=splitName[0]+'.xlsx')
        targetInputFile = splitName[0]+'.xlsx'
    elif(splitName[1]=='lsx'):
        print('Its a xlsx file')
        targetInputFile = fileName
    else:
        print('Invalid file format! Use either .xls or .xlsx')
        sys.exit()

    # Input Excel
    print("Loading Input file: " + targetInputFile)
    wb = openpyxl.load_workbook(targetInputFile)
    # Get Correct Sheet
    listOfSheets = wb.get_sheet_names()
    # print(listOfSheets)
    countOfSheets = 0
    for sheet in listOfSheets:
        print('Press '+str(countOfSheets)+' for sheet '+sheet)
        countOfSheets+=1
    choice = input('Enter Your Choice: ')
    # sh1 = wb.get_active_sheet()
    print('Sheet Selected: '+listOfSheets[int(choice)])
    sh1 = wb.get_sheet_by_name(listOfSheets[int(choice)])
    maxRows = sh1.max_row
    # Output Excel
    wb_O = openpyxl.Workbook()
    # sh1_O = wb_O.get_active_sheet()
    sh1_O = wb_O.create_sheet(index=0, title='-Current')
    sh2_O = wb_O.create_sheet(index=1, title='+Current')
    sh3_O = wb_O.create_sheet(index=2, title='Charge_Discharge')
    outputFile = 'output'
    outputFile = input("Enter a Name for Output File: ")
    rowStarts = input('Row Starts at 2. Press Enter to confirm or Type a New value & Hit Enter')
    if(rowStarts == '' or rowStarts==None):
        rowStart=2
    else:
        rowStart = int(rowStarts)
    version = 0.1
    groupLength = []
    columnStart = 'A'
    columnStart3 = 'A'
    index = 1
    sh3_index = 1
    prevNegative = 'N'
    potential_col = input("Potential: ")
    current_col = input("Current: ")
    time_col = input("Time: ")
    capacity_col = input("Capacity: ")
    chdc_gap = input("Gap between Charge-Discharge: ")

    @staticmethod
    def negativeProc(current,i):
        AutoExcel.prevNegative = 'Y'
        potentialCoordinate = AutoExcel.columnStart + str(AutoExcel.index)
        currentCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 1) + str(AutoExcel.index)
        timeCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 2) + str(AutoExcel.index)
        capacityCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 3) + str(AutoExcel.index)

        print("the coordinates are: " + potentialCoordinate + " " + currentCoord + " " + capacityCoord)
        potential = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.potential_col))
        time = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.time_col))
        capacity = AutoExcel.sh1.cell(row=i, column=column_index_from_string(AutoExcel.capacity_col))

        AutoExcel.sh1_O[potentialCoordinate] = potential.value
        AutoExcel.sh1_O[currentCoord] = current.value
        AutoExcel.sh1_O[timeCoord] = time.value
        AutoExcel.sh1_O[capacityCoord] = capacity.value
        AutoExcel.index += 1
        print("Row#: " + str(i) + "  Saving " + str(current.value) + " in output..")
        if i == AutoExcel.maxRows:
            AutoExcel.groupLength.append(AutoExcel.index - 1)
            AutoExcel.capture_Ch_Dc(i+1)

    @staticmethod
    def capture_Ch_Dc(i):
        if AutoExcel.chdc_gap==None or AutoExcel.chdc_gap==" " or AutoExcel.chdc_gap=="/n":
            AutoExcel.chdc_gap = 0
        lastIndex = i-1-int(AutoExcel.chdc_gap)
        # print("last index: "+ str(lastIndex))
        # print("sh3_index: "+ str(AutoExcel.sh3_index))
        potentialCoordinate3 = AutoExcel.columnStart3 + str(AutoExcel.sh3_index)
        currentCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 1) + str(AutoExcel.sh3_index)
        timeCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 2) + str(AutoExcel.sh3_index)
        capacityCoord3 = get_column_letter(int(column_index_from_string(AutoExcel.columnStart3)) + 3) + str(AutoExcel.sh3_index)
        print("the coordinates are: " + potentialCoordinate3 + " " + currentCoord3 + " " +timeCoord3 + " "+capacityCoord3)

        potential3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.potential_col))
        current3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.current_col))
        time3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.time_col))
        capacity3 = AutoExcel.sh1.cell(row=lastIndex, column=column_index_from_string(AutoExcel.capacity_col))

        AutoExcel.sh3_O[potentialCoordinate3] = potential3.value
        AutoExcel.sh3_O[currentCoord3] = current3.value
        AutoExcel.sh3_O[timeCoord3] = time3.value
        AutoExcel.sh3_O[capacityCoord3] = capacity3.value
        AutoExcel.sh3_index += 1

    @staticmethod
    def addChart(groupLength):
        print('Printing Chart...')
        chart = ScatterChart()
        chart.title = "Capacity Vs Voltage"
        chart.style = 13
        chart.x_axis.title = 'Capacity'
        chart.y_axis.title = 'Voltage'
        le = (len(groupLength) * 4) + len(groupLength)
        print("The sizes of each grp" + str(groupLength[0]))
        k = 0
        m = 1
        # print("list siz = " + str(len(groupLength)))
        # print("le = " + str(le))
        for j in range(4, le, 5):
            # print("j = " + str(j))
            xvalues = Reference(AutoExcel.sh1_O, min_col=j, min_row=1, max_row=groupLength[k])
            values = Reference(AutoExcel.sh1_O, min_col=m, min_row=1, max_row=groupLength[k])
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)
            k += 1
            m += 5

        AutoExcel.sh1_O.add_chart(chart, 'L3')

        AutoExcel.wb_O.save(AutoExcel.outputFile+'.xlsx')

    @staticmethod
    def mainApp():

        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        for i in range(AutoExcel.rowStart, AutoExcel.maxRows+1):
            current = AutoExcel.sh1.cell(row=i,column=column_index_from_string(AutoExcel.current_col))
            if type(AutoExcel.columnStart) == type(current.value):
                currentVal = float(current.value)
            elif type(AutoExcel.index) == type(current.value) or type(AutoExcel.version) == type(current.value):
                currentVal = current.value
            else:
                pass
                # print("Invalid Current Value Type: "+str(type(current.value)))
            if current.value == None:
                current.value = 0
            if currentVal<0:
                if AutoExcel.prevNegative == 'N':
                   AutoExcel.capture_Ch_Dc(i)
                AutoExcel.negativeProc(current,i)
            elif AutoExcel.prevNegative == 'Y':
                AutoExcel.capture_Ch_Dc(i)
                AutoExcel.prevNegative = 'N'
                potentialCoordinate = AutoExcel.columnStart + str(AutoExcel.index)
                currentCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 1) + str(AutoExcel.index)
                timeCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 2) + str(AutoExcel.index)
                capacityCoord = get_column_letter(int(column_index_from_string(AutoExcel.columnStart)) + 3) + str(AutoExcel.index)
                AutoExcel.sh1_O[currentCoord].fill = redFill
                AutoExcel.sh1_O[potentialCoordinate].fill = redFill
                AutoExcel.sh1_O[timeCoord].fill = redFill
                AutoExcel.sh1_O[capacityCoord].fill = redFill
                AutoExcel.groupLength.append(AutoExcel.index-1)
                AutoExcel.index = 1
                AutoExcel.columnStart = get_column_letter(int(column_index_from_string(AutoExcel.columnStart))+5)
            else:
                continue
        print("Saving values..")
        AutoExcel.wb_O.save(AutoExcel.outputFile+'.xlsx')

        AutoExcel.addChart(AutoExcel.groupLength)

        print("xxxxx------ Program Ended ------xxx")
        print("Output File Created: "+AutoExcel.outputFile+".xlsx")

AutoExcel.mainApp()